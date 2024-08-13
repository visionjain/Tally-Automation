import pyautogui
import pygetwindow as gw
import openpyxl
import time
from datetime import datetime
from dateutil import parser
import tkinter as tk
from tkinter import filedialog, simpledialog, ttk
import win32gui
import win32con

def bring_window_to_front(window_title):
    """ Bring the specified window to the foreground. """
    try:
        hwnd = win32gui.FindWindow(None, window_title)
        if hwnd:
            win32gui.ShowWindow(hwnd, win32con.SW_RESTORE)
            win32gui.SetForegroundWindow(hwnd)
    except Exception as e:
        print(f"Error bringing window to front: {e}")

def parse_amount(amount):
    """ Convert amount to float, removing any commas and handling various formats. """
    try:
        if isinstance(amount, str):
            return float(amount.replace(',', '').replace('₹', '').replace('$', '').strip())
        elif isinstance(amount, (int, float)):
            return float(amount)
        else:
            return 0
    except (ValueError, AttributeError):
        return 0

def process_excel(file_path, bank_name, progress_var, progress_bar, file_label, completed_label, failed_label):
    try:
        wb = openpyxl.load_workbook(file_path)
        sheet = wb.active

        # Find the TallyPrime window
        tally_window = None
        for window in gw.getAllTitles():
            if "TallyPrime" in window:
                tally_window = window
                break

        if tally_window:
            tally_win = gw.getWindowsWithTitle(tally_window)[0]
            tally_win.activate()
            time.sleep(1)
            bring_window_to_front(tally_window)  # Bring the Tkinter window to the foreground

            total_rows = sheet.max_row - 1  # Number of rows excluding the header
            progress_var.set(0)
            progress_bar['maximum'] = total_rows

            completed_rows = 0
            failed_rows = 0

            for row in range(2, sheet.max_row + 1):  # Start from row 2 to skip header
                first_date = sheet[f'A{row}'].value
                withdrawal_amount = sheet[f'B{row}'].value
                deposit_amount = sheet[f'C{row}'].value

                print(f"Processing row {row}: Date={first_date}, Withdrawal Amount={withdrawal_amount}, Deposit Amount={deposit_amount}")

                if isinstance(first_date, (datetime, str)):
                    try:
                        parsed_date = parser.parse(str(first_date))
                        formatted_date = parsed_date.strftime('%d-%m-%y')
                    except ValueError:
                        print(f"Error in row {row}: Date format is incorrect. Content={first_date}")
                        failed_rows += 1
                        progress_var.set(row - 1)  # Update to the current row number
                        completed_label.config(text=f"Completed: {completed_rows}")
                        failed_label.config(text=f"Failed: {failed_rows}")
                        root.update_idletasks()
                        continue
                else:
                    print(f"Error in row {row}: Date format is incorrect or not a datetime object. Content={first_date}")
                    failed_rows += 1
                    progress_var.set(row - 1)  # Update to the current row number
                    completed_label.config(text=f"Completed: {completed_rows}")
                    failed_label.config(text=f"Failed: {failed_rows}")
                    root.update_idletasks()
                    continue

                withdrawal_amount = parse_amount(withdrawal_amount)
                deposit_amount = parse_amount(deposit_amount)

                print(f"Converted amounts: Withdrawal Amount={withdrawal_amount}, Deposit Amount={deposit_amount}")

                if row == 2:
                    pyautogui.press('v')

                pyautogui.press('f2')
                pyautogui.write(formatted_date)
                time.sleep(0.5)
                pyautogui.press('enter')
                if withdrawal_amount > 0:
                    pyautogui.press('f5')
                elif deposit_amount > 0:
                    pyautogui.press('f6')
                else:
                    print(f"Error in row {row}: No valid amount found. Withdrawal Amount={withdrawal_amount}, Deposit Amount={deposit_amount}")
                    failed_rows += 1
                    progress_var.set(row - 1)  # Update to the current row number
                    completed_label.config(text=f"Completed: {completed_rows}")
                    failed_label.config(text=f"Failed: {failed_rows}")
                    root.update_idletasks()
                    continue

                pyautogui.write(bank_name)
                time.sleep(0.5)
                pyautogui.press('enter')
                pyautogui.write('Suspense')
                time.sleep(0.5)
                pyautogui.press('enter')

                if withdrawal_amount > 0:
                    pyautogui.write(f"{withdrawal_amount}")
                elif deposit_amount > 0:
                    pyautogui.write(f"{deposit_amount}")

                pyautogui.hotkey('ctrl', 'a')
                time.sleep(0.5)

                completed_rows += 1
                progress_var.set(row - 1)  # Update to the current row number
                completed_label.config(text=f"Completed: {completed_rows}")
                failed_label.config(text=f"Failed: {failed_rows}")
                root.update_idletasks()

            print("Processing completed for all rows.")
        else:
            print("TallyPrime application window not found.")
    except FileNotFoundError:
        print("The specified file was not found.")

def browse_file(file_path_var, file_label):
    """ Open file dialog to select an Excel file and update the UI. """
    file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
    if file_path:
        file_path_var.set(file_path)
        file_label.config(text=f"Selected file: {file_path}")

def main():
    global root

    root = tk.Tk()
    root.title("Tally Data Entry")

    # Make the Tkinter window always on top
    root.attributes('-topmost', True)

    tk.Label(root, text="Select Excel file:").pack(pady=5)

    file_path_var = tk.StringVar()
    file_path_var.set("No file selected")
    file_label = tk.Label(root, textvariable=file_path_var, wraplength=400)
    file_label.pack(pady=5)

    file_button = tk.Button(root, text="Browse", command=lambda: browse_file(file_path_var, file_label))
    file_button.pack(pady=5)

    tk.Label(root, text="Bank Name:").pack(pady=5)
    bank_name_entry = tk.Entry(root)
    bank_name_entry.pack(pady=5)

    start_button = tk.Button(root, text="Start", command=lambda: process_excel(file_path_var.get(), bank_name_entry.get(), progress_var, progress_bar, file_label, completed_label, failed_label))
    start_button.pack(pady=5)

    progress_var = tk.IntVar()
    progress_bar = ttk.Progressbar(root, length=400, maximum=1, variable=progress_var)
    progress_bar.pack(pady=20)

    # Labels to show completed and failed rows
    completed_label = tk.Label(root, text="Completed: 0")
    completed_label.pack(pady=2)

    failed_label = tk.Label(root, text="Failed: 0")
    failed_label.pack(pady=2)

    root.mainloop()

if __name__ == "__main__":
    main()
